from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

from audits.jake_audit_workbook import (
    AuditRow,
    LiveContext,
    _apply_live_status,
    _build_controller_verification,
    _known_mac_bug_kind,
    _row_state,
    _weighted_ready_score,
    generate_nycha_audit_workbook,
)
from mcp import jake_ops_mcp as opsmod


ROOT = Path(__file__).resolve().parents[1]
FIXTURE_DIR = ROOT / "tests" / "fixtures" / "audit"
BASELINE_JSON = ROOT / "tests" / "baselines" / "audit_baseline.json"
BASELINE_WORKBOOK = ROOT / "tests" / "baselines" / "audit_fixture_workbook.xlsx"


def _configure_audit_fixture_env(monkeypatch) -> None:
    monkeypatch.setenv("JAKE_AUDIT_TEMPLATE_WORKBOOK", str(FIXTURE_DIR / "nycha_template.xlsx"))
    monkeypatch.setenv("JAKE_NYCHA_INFO_CSV", str(FIXTURE_DIR / "nycha_info_fixture.csv"))
    monkeypatch.setenv("JAKE_TAUC_AUDIT_CSV", str(FIXTURE_DIR / "tauc_nycha_audit_fixture.csv"))
    monkeypatch.setenv("JAKE_VILO_INVENTORY_SNAPSHOT", str(FIXTURE_DIR / "vilo_inventory_fixture.json"))
    opsmod.NYCHA_INFO_CSV = Path(str(FIXTURE_DIR / "nycha_info_fixture.csv"))
    opsmod.TAUC_NYCHA_AUDIT_CSV = Path(str(FIXTURE_DIR / "tauc_nycha_audit_fixture.csv"))
    opsmod.load_nycha_info_rows.cache_clear()
    opsmod.load_tauc_nycha_audit_rows.cache_clear()


def _load_baseline() -> dict[str, object]:
    return json.loads(BASELINE_JSON.read_text(encoding="utf-8"))


def _make_live_context() -> LiveContext:
    switch_identity = "000007.001.SW01"
    return LiveContext(
        building_id="000007.001",
        site_id="000007",
        online_units_by_token={},
        exact_matches_by_unit={},
        active_alert_count=0,
        building_device_count=7,
        site_online_count=7,
        inferred_switch_identity=switch_identity,
        live_port_macs_by_interface={
            "ether1": ["aa:11:22:33:44:55"],
            "ether2": ["11:22:33:44:55:66"],
            "ether3": ["20:22:33:44:55:61"],
            "ether4": ["aa:11:22:33:44:55"],
            "ether5": ["de:ad:be:ef:00:01"],
            "ether6": [],
            "ether7": ["aa:11:22:33:44:55"],
        },
        switch_identities_by_label_prefix={"SW1": switch_identity},
        live_port_macs_by_switch_identity={
            switch_identity: {
                "ether1": ["aa:11:22:33:44:55"],
                "ether2": ["11:22:33:44:55:66"],
                "ether3": ["20:22:33:44:55:61"],
                "ether4": ["aa:11:22:33:44:55"],
                "ether5": ["de:ad:be:ef:00:01"],
                "ether6": [],
                "ether7": ["aa:11:22:33:44:55"],
            }
        },
        controller_verification_by_mac={
            "aa:11:22:33:44:55": {"status": "match", "label": "Vilo inventory match"},
            "10:22:33:44:55:66": {"status": "match", "label": "TAUC audit match"},
            "20:22:33:44:55:60": {"status": "match", "label": "TAUC audit match"},
            "30:22:33:44:55:70": {"status": "match", "label": "Vilo inventory match"},
            "40:22:33:44:55:80": {"status": "unverified", "label": "Not verified"},
            "50:22:33:44:55:90": {"status": "mismatch", "label": "TAUC audit mismatch"},
        },
        live_failures=[],
    )


def _failure_live_context() -> LiveContext:
    return LiveContext(
        building_id="000007.001",
        site_id="000007",
        online_units_by_token={},
        exact_matches_by_unit={},
        active_alert_count=0,
        building_device_count=0,
        site_online_count=0,
        inferred_switch_identity="000007.001.SW01",
        live_port_macs_by_interface={},
        switch_identities_by_label_prefix={"SW1": "000007.001.SW01"},
        live_port_macs_by_switch_identity={},
        controller_verification_by_mac={},
        live_failures=[
            {
                "source": "get_site_summary",
                "classification": "missing_runtime",
                "detail": "get_site_summary timed out after 3.0s",
            }
        ],
    )


def _row_case(unit_label: str, mac: str, switch_port: str, pppoe_unit: str | None = None) -> AuditRow:
    return AuditRow(
        unit_key=unit_label,
        unit_label=unit_label,
        mac_cpe=mac,
        pppoe_unit=pppoe_unit or unit_label,
        notes="Good",
        image_ap_make="",
        image_ap_sticker_apartment="",
        image_ap_mac="",
        inventory_mac_verification="",
        implication="",
        action="None",
        switch_port=switch_port,
    )


def _source_row(unit_label: str, mac: str, ap_make: str, serial: str, pppoe_unit: str | None = None) -> dict[str, str]:
    return {
        "Unit": unit_label,
        "PPPoE": f"site-{pppoe_unit or unit_label}",
        "MAC Address": mac,
        "AP Make": ap_make,
        "AP Serial Number": serial,
    }


def _sheet_values(path: Path) -> list[list[object]]:
    wb = load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    values: list[list[object]] = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True):
        values.append(list(row))
    return values


def _normalized_sheet_values(path: Path) -> list[list[object]]:
    values = _sheet_values(path)
    if values and len(values[0]) >= 6:
        values[0][5] = "29% Verified Ready"
    return values


def test_audit_baseline_matches_reviewed_reference_with_green_only_readiness(tmp_path, monkeypatch) -> None:
    _configure_audit_fixture_env(monkeypatch)
    baseline = _load_baseline()
    output_path = tmp_path / "audit_fixture_workbook.xlsx"
    workbook_result = generate_nycha_audit_workbook(
        "123-125 Test St",
        out_path=output_path,
        template_path=FIXTURE_DIR / "nycha_template.xlsx",
        ops=None,
        _live_context_override=_make_live_context(),
    )

    expected_workbook = baseline["workbook_result"]
    assert workbook_result["row_count"] == expected_workbook["row_count"]
    assert workbook_result["weighted_ready_percent"] == 29
    assert workbook_result["rows"] == expected_workbook["rows"]
    assert any(row["notes"] == "WRONG UNIT" for row in workbook_result["rows"])
    assert _normalized_sheet_values(output_path) == _normalized_sheet_values(BASELINE_WORKBOOK)


def test_audit_contract_states_match_reviewed_baseline(monkeypatch) -> None:
    _configure_audit_fixture_env(monkeypatch)
    baseline = _load_baseline()
    live = _make_live_context()

    cases = [
        ("1A", "AA:11:22:33:44:55", "Vilo", "VILO-1A", "SW1-1", None),
        ("1B", "10:22:33:44:55:66", "TP-Link", "HC220-1B", "SW1-2", None),
        ("1C", "20:22:33:44:55:60", "TP-Link", "HC220-1C", "SW1-3", None),
        ("1D", "30:22:33:44:55:70", "Vilo", "VILO-1D", "SW1-4", None),
        ("1E", "40:22:33:44:55:80", "Vilo", "VILO-1E", "SW1-5", None),
        ("1F", "50:22:33:44:55:90", "TP-Link", "HC220-1F", "SW1-6", None),
        ("1G", "AA:11:22:33:44:55", "Vilo", "VILO-1A", "SW1-7", "1A"),
    ]
    evaluated_rows = []
    for unit, mac, vendor, serial, port, pppoe_unit in cases:
        evaluated = _apply_live_status(_row_case(unit, mac, port, pppoe_unit), _source_row(unit, mac, vendor, serial, pppoe_unit), live)
        evaluated_rows.append(
            {
                "unit": unit,
                "notes": evaluated.notes,
                "inventory_mac_verification": evaluated.inventory_mac_verification,
                "state": _row_state(evaluated),
                "mac_cpe": evaluated.mac_cpe,
                "implication": evaluated.implication,
                "action": evaluated.action,
            }
        )

    failure_row = _apply_live_status(
        _row_case("1Z", "60:22:33:44:55:90", "SW1-8"),
        _source_row("1Z", "60:22:33:44:55:90", "TP-Link", "HC220-1Z"),
        _failure_live_context(),
    )
    monkeypatch.setenv("JAKE_VILO_INVENTORY_SNAPSHOT", str(FIXTURE_DIR / "vilo_inventory_fixture_bad.json"))
    _, vilo_failures = _build_controller_verification([_source_row("1A", "AA:11:22:33:44:55", "Vilo", "VILO-1A")])

    assert {
        "exact": _known_mac_bug_kind("aa:11:22:33:44:55", "aa:11:22:33:44:55"),
        "first_octet": _known_mac_bug_kind("10:22:33:44:55:66", "11:22:33:44:55:66"),
        "last_octet": _known_mac_bug_kind("20:22:33:44:55:60", "20:22:33:44:55:61"),
        "no_match": _known_mac_bug_kind("20:22:33:44:55:60", "20:22:33:44:55:62"),
    } == baseline["known_mac_bug_kind"]
    assert evaluated_rows == baseline["evaluated_rows"]
    assert {
        "notes": failure_row.notes,
        "state": _row_state(failure_row),
        "implication": failure_row.implication,
        "action": failure_row.action,
    } == baseline["failure_row"]
    assert vilo_failures == baseline["vilo_snapshot_failure"]


def test_bug_adjusted_match_uses_observed_mac_as_legacy_mac_cpe(monkeypatch) -> None:
    _configure_audit_fixture_env(monkeypatch)
    live = _make_live_context()

    evaluated = _apply_live_status(
        _row_case("1B", "10:22:33:44:55:66", "SW1-2"),
        _source_row("1B", "10:22:33:44:55:66", "TP-Link", "HC220-1B"),
        live,
    )

    assert evaluated.notes == "Good"
    assert evaluated.inventory_mac_verification == "Bug-adjusted match"
    assert evaluated.mac_cpe == "11:22:33:44:55:66"


def test_weighted_ready_excludes_yellow_rows() -> None:
    rows = [
        AuditRow("1A", "1A", "", "1A", "Good", "", "", "", "", "", ""),
        AuditRow("1B", "1B", "", "1B", "Good", "", "", "", "", "", ""),
        AuditRow("1C", "1C", "", "1C", "MOVE CPE TO WAN PORT", "", "", "", "", "", ""),
        AuditRow("1D", "1D", "", "1D", "MOVE CPE TO CORRECT UNIT", "", "", "", "", "", ""),
        AuditRow("1E", "1E", "", "1E", "UNKNOWN MAC ON PORT", "", "", "", "", "", ""),
        AuditRow("1F", "1F", "", "1F", "CONTROLLER MISMATCH", "", "", "", "", "", ""),
        AuditRow("1G", "1G", "", "1G", "WRONG UNIT", "", "", "", "", "", ""),
    ]

    assert _weighted_ready_score(rows) == 29


def test_yellow_rows_contribute_zero_to_verified_ready() -> None:
    rows = [
        AuditRow("1C", "1C", "", "1C", "MOVE CPE TO WAN PORT", "", "", "", "", "", ""),
        AuditRow("1D", "1D", "", "1D", "MOVE CPE TO CORRECT UNIT", "", "", "", "", "", ""),
        AuditRow("1E", "1E", "", "1E", "UNKNOWN MAC ON PORT", "", "", "", "", "", ""),
    ]

    assert _weighted_ready_score(rows) == 0
